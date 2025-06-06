import openpyxl
import json
import re
from openpyxl.styles import Border, Side, Font, PatternFill, Alignment, GradientFill
from openpyxl.chart import BarChart, PieChart, Reference, DoughnutChart
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import DataBarRule, ColorScaleRule
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import os
import subprocess
import platform
from datetime import datetime

class TimeTableProcessor:
    def __init__(self):
        self.setup_gui()
        
    def setup_gui(self):
        self.root = tk.Tk()
        self.root.title("시수배정현황 처리 프로그램")
        self.root.geometry("800x800")

        # ttk 스타일 설정
        style = ttk.Style()
        style.configure('TLabelframe', padding=5)
        style.configure('TButton', padding=5)
        
        # 메인 프레임
        main_frame = ttk.Frame(self.root, padding=10)
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # 파일 선택 영역
        file_frame = ttk.LabelFrame(main_frame, text="파일 선택")
        file_frame.pack(fill=tk.X, padx=5, pady=5)
        
        # 진행 상태바 추가
        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(file_frame,
                                          variable=self.progress_var,
                                          maximum=100,
                                          mode='determinate')
        self.progress_bar.pack(fill=tk.X, padx=5, pady=5)
        
        # 파일 목록 프레임
        file_list_frame = ttk.Frame(file_frame)
        file_list_frame.pack(fill=tk.BOTH, expand=True)
        
        # 파일 목록 Text 위젯과 스크롤바
        self.file_text = tk.Text(file_list_frame, height=6, wrap=tk.WORD)
        file_scrollbar = ttk.Scrollbar(file_list_frame, command=self.file_text.yview)
        
        self.file_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(5,0))
        file_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.file_text.configure(yscrollcommand=file_scrollbar.set)
        
        # 파일 선택 버튼 프레임
        button_frame = ttk.Frame(file_frame)
        button_frame.pack(fill=tk.X, pady=5)
        
        select_button = ttk.Button(button_frame, text="파일 선택", command=self.select_files)
        select_button.pack(side=tk.LEFT, padx=5)
        
        clear_button = ttk.Button(button_frame, text="선택 초기화", command=self.clear_selection)
        clear_button.pack(side=tk.LEFT, padx=5)
        
        # 상태 표시 영역
        status_frame = ttk.LabelFrame(main_frame, text="처리 상태")
        status_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # 상태 Text 위젯과 스크롤바
        self.status_text = tk.Text(status_frame, wrap=tk.WORD)
        status_scrollbar = ttk.Scrollbar(status_frame, command=self.status_text.yview)
        
        self.status_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(5,0), pady=5)
        status_scrollbar.pack(side=tk.RIGHT, fill=tk.Y, pady=5)
        self.status_text.configure(yscrollcommand=status_scrollbar.set)

        # 통계 모드 선택 영역
        option_frame = ttk.LabelFrame(main_frame, text="통계 모드")
        option_frame.pack(fill=tk.X, padx=5, pady=5)

        self.mode_var = tk.StringVar(value="single")
        single_radio = ttk.Radiobutton(option_frame,
                                       text="단일 학교 (파일 합침)",
                                       variable=self.mode_var,
                                       value="single")
        multi_radio = ttk.Radiobutton(option_frame,
                                      text="학교별 통계",
                                      variable=self.mode_var,
                                      value="multi")
        single_radio.pack(side=tk.LEFT, padx=5)
        multi_radio.pack(side=tk.LEFT, padx=5)

        # 하단 버튼 영역
        bottom_frame = ttk.Frame(main_frame)
        bottom_frame.pack(fill=tk.X, padx=5, pady=5)
        
        # 자동 열기 체크박스
        self.auto_open_var = tk.BooleanVar(value=True)
        auto_open_check = ttk.Checkbutton(bottom_frame,
                                        text="처리 후 자동으로 파일 열기",
                                        variable=self.auto_open_var)
        auto_open_check.pack(side=tk.LEFT, padx=5)
        
        # 처리 시작 버튼
        process_button = ttk.Button(bottom_frame, text="처리 시작", command=self.process_files)
        process_button.pack(side=tk.RIGHT, padx=5)
        
        # 종료 버튼
        quit_button = ttk.Button(bottom_frame, text="종료", command=self.root.quit)
        quit_button.pack(side=tk.RIGHT, padx=5)
    
    def normalize_subject_name(self, subject_name):
        """과목명 정규화 함수 - 숫자와 특수문자 제거하여 매칭용 키 생성"""
        if not subject_name:
            return subject_name
            
        # 1. '*' 문자 제거
        normalized = subject_name.lstrip('*').strip()
        
        # 2. 끝에 붙은 숫자 제거 (예: "수학1" -> "수학", "영어2" -> "영어")
        normalized = re.sub(r'\d+$', '', normalized).strip()
        
        # 3. 로마숫자 제거 (예: "수학Ⅰ" -> "수학", "영어Ⅱ" -> "영어")
        normalized = re.sub(r'[ⅠⅡⅢⅣⅤⅥⅦⅧⅨⅩ]+$', '', normalized).strip()
        
        # 4. 괄호와 그 안의 내용 제거 (예: "수학(미적분)" -> "수학")
        normalized = re.sub(r'\([^)]*\)$', '', normalized).strip()
        
        # 5. 레벨 표시 제거 (예: "수학 A" -> "수학", "영어 고급" -> "영어")
        normalized = re.sub(r'\s+(A|B|C|고급|중급|초급|기초|심화)$', '', normalized).strip()
        
        # 디버깅용 로그
        if normalized != subject_name.lstrip('*').strip():
            print(f"과목명 정규화: '{subject_name}' -> '{normalized}'")
            
        return normalized
    
    def get_subject_group(self, subject_name, subject_group_mapping):
        """과목명으로 교과(군) 찾기 - 정규화된 이름으로 매칭 시도"""
        if not subject_name:
            return '기타'
            
        # 1. 원본 이름으로 먼저 시도
        original_key = subject_name.lstrip('*')
        if original_key in subject_group_mapping:
            return subject_group_mapping[original_key]
        
        # 2. 정규화된 이름으로 시도
        normalized_key = self.normalize_subject_name(subject_name)
        if normalized_key in subject_group_mapping:
            print(f"정규화된 매칭 성공: '{subject_name}' -> '{normalized_key}' -> '{subject_group_mapping[normalized_key]}'")
            return subject_group_mapping[normalized_key]
        
        # 3. 부분 매칭 시도 (정규화된 이름이 매핑 키에 포함되어 있는지)
        for key, group in subject_group_mapping.items():
            if normalized_key in key or key in normalized_key:
                print(f"부분 매칭 성공: '{subject_name}' -> '{key}' -> '{group}'")
                return group
        
        # 4. 매칭 실패시 기타로 분류하고 로그 출력
        print(f"매칭 실패: '{subject_name}' (정규화: '{normalized_key}') -> '기타'")
        return '기타'

    def clear_selection(self):
        """선택된 파일 목록 초기화"""
        self.file_text.delete('1.0', tk.END)
        if hasattr(self, 'file_paths'):
            del self.file_paths
        self.add_log("파일 선택이 초기화되었습니다.")
        self.update_progress(0)

    def update_progress(self, value, message=""):
        """진행 상태바 업데이트"""
        self.progress_var.set(value)
        if message:
            self.add_log(message)
        self.root.update()

    def add_log(self, message):
        """로그 메시지 추가 (타임스탬프 포함)"""
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        self.status_text.insert(tk.END, f"[{timestamp}] {message}\n")
        self.status_text.see(tk.END)
        self.root.update()

    def autofit_columns(self, ws, min_width=10, padding=2):
        """워크시트의 열 너비를 내용에 맞게 자동 조정"""
        for column_cells in ws.columns:
            max_length = 0
            for cell in column_cells:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = max(max_length + padding, min_width)
            ws.column_dimensions[column_cells[0].column_letter].width = adjusted_width

    def open_file(self, path):
        """Save 작업 후 파일을 여는 OS별 함수"""
        if not self.auto_open_var.get():
            return
        system = platform.system()
        try:
            if system == 'Windows':
                os.startfile(path)
            elif system == 'Darwin':
                subprocess.Popen(['open', path])
            else:
                subprocess.Popen(['xdg-open', path])
        except Exception as e:
            self.add_log(f"파일 자동 열기에 실패했습니다: {e}")
    
    def filter_subject_groups(self, subject_groups):
        """교과 그룹 필터링 규칙"""
        # 디버깅용 출력 추가
        print("처리 전 교과군:", subject_groups)
        
        subject_groups = set(subject_groups)  # 중복 제거
        filtered = subject_groups.copy()

        # 정확한 교과명으로 규칙 정의
        rules = {
            frozenset(['과학', '기술·가정']): '과학',  
            frozenset(['과학', '기술∙가정']): '과학',  
            frozenset(['보건', '기술·가정']): '기술∙가정',  
            frozenset(['제2외국어', '한문']): '제2외국어',
            frozenset(['수학', '정보']): '수학',
            frozenset(['과학', '보건']): '과학',
            frozenset(['한문', '보건']): '한문',
            frozenset(['국어', '한문']): '국어',
            frozenset(['영어', '예술']): '영어',
            frozenset(['사회', '예술']): '사회',
            frozenset(['과학', '정보']): '과학',
            frozenset(['예술', '제2외국어']): '예술',
            frozenset(['정보', '기술·가정']): '정보',
            frozenset(['수학', '기술·가정']): '수학', 
            frozenset(['영어', '기술·가정']): '영어', 
            frozenset(['국어', '보건']): '국어',  # 추가
            frozenset(['수학', '보건']): '수학',  # 추가
            frozenset(['영어', '보건']): '영어',  # 추가
            frozenset(['예술', '보건']): '예술',  # 추가
            frozenset(['국어', '전문 교과']): '국어',  # 추가
            frozenset(['수학', '전문 교과']): '수학',  # 추가
            frozenset(['영어', '전문 교과']): '영어',  # 추가
            frozenset(['예술', '전문 교과']): '예술',  # 추가
            frozenset(['진로', '전문 교과']): '전문 교과',  # 추가
            frozenset(['제2외국어', '보건']): '제2외국어',  # 추가
        }

        # 교양, 기타 처리
        other_subjects = {subj for subj in filtered if subj not in ['교양', '기타']}
        if other_subjects:
            filtered = other_subjects

        # 규칙 적용
        current_subjects = frozenset(filtered)
        for rule_set, result in rules.items():
            if rule_set & current_subjects == rule_set:  # 교집합이 규칙 집합과 같으면
                filtered = {result}
                print(f"규칙 적용됨: {rule_set} -> {result}")
                break

        # 전문 교과, 진로 처리
        if any(x in filtered for x in ['전문 교과', '진로']):
            other_subjects = {x for x in filtered if x not in ['전문 교과', '진로']}
            if other_subjects:
                filtered = other_subjects
                print("전문 교과/진로 규칙 적용됨")

        result = sorted(filtered)
        print("처리 후 교과군:", result)
        return result
    
    def select_files(self):
        file_paths = filedialog.askopenfilenames(
            title="시수배정현황 파일 선택",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        
        if file_paths:
            self.file_paths = file_paths
            # Text 위젯 내용 초기화
            self.file_text.delete('1.0', tk.END)
            # 선택된 파일 목록 표시
            self.file_text.insert(tk.END, "선택된 파일:\n")
            for path in file_paths:
                self.file_text.insert(tk.END, f"• {os.path.basename(path)}\n")
            self.add_log("파일이 선택되었습니다.")

    def load_subject_group_mapping(self, json_path):
        """JSON 파일에서 교과(군) 모집 데이터를 불러오는 함수"""
        try:
            # BOM 제거를 위해 utf-8-sig 사용
            with open(json_path, 'r', encoding='utf-8-sig') as f:
                data = json.load(f)
            print(f"[매핑 로드] 총 {len(data)}개 키를 로드했습니다. (예시: {list(data.keys())[:5]})")
            return data
        except FileNotFoundError:
            print("교과(군) 매핑 파일을 찾을 수 없습니다.")
            messagebox.showerror("Error", "교과(군) 매핑 파일을 찾을 수 없습니다.")
            return {}

    def extract_data(self, ws):
        """워크시트에서 데이터를 추출하는 함수"""
        results = []
        total_hours_col = None
        is_header_found = False
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row), 1):
            values = [cell.value for cell in row]
            
            # 빈 행 건너뛰기
            if not any(values):
                continue
            
            # 총시수 열 찾기
            if "총시수" in values:
                total_hours_col = values.index("총시수")
                is_header_found = True
                continue
            
            # 헤더를 찾은 후에만 데이터 처리
            if is_header_found and len(values) > 3:  # 최소 D열까지는 있어야 함
                if values[1] and isinstance(values[1], str):  # B열 체크 (과목)
                    if values[1] != "과목" and not str(values[1]).startswith('20'):
                        # 총시수 확인
                        if total_hours_col is not None and len(values) > total_hours_col:
                            total_hours = values[total_hours_col]
                            if isinstance(total_hours, (int, float)):
                                # 과목(B열)과 교사명(D열) 추출
                                subject_full = values[1].strip()
                                
                                # 과목명에서 첫 번째 하이픈 전까지만 추출
                                subject = subject_full.split('-')[0].strip()
                                
                                teacher = values[3].strip() if values[3] else ""  # D열이 교사명
                                
                                anonymized_teacher = teacher[0] + '*' * (len(teacher) - 2) + teacher[-1] if len(teacher) > 1 else teacher
                                self.add_log(f"데이터 발견: {subject} (원본: {subject_full}) - {anonymized_teacher} - {total_hours}")
                                
                                results.append({
                                    '과목': subject,
                                    '교사명': teacher,
                                    '총시수': int(total_hours)
                                })
        return results

    def process_workbook(self, wb):
        """워크북 전체 처리"""
        all_results = []
        
        for ws in wb.worksheets:
            results = self.extract_data(ws)
            if results:
                all_results.extend(results)
                
        # 중복 제거 및 데이터 변환
        merged_results = {}
        for item in all_results:
            key = (item['과목'], item['교사명'])
            if key not in merged_results:
                merged_results[key] = item
            else:
                # 같은 과목-교사 조합이 있으면 시수 확인하여 큰 값 사용
                if item['총시수'] > merged_results[key]['총시수']:
                    merged_results[key] = item
        
        return list(merged_results.values())

    def process_files(self):
        if not hasattr(self, 'file_paths'):
            messagebox.showwarning("경고", "먼저 파일을 선택해주세요.")
            return

        try:
            self.add_log("파일 처리 시작...")
            school_data = []  # 각 학교별 데이터를 저장할 리스트
            school_names = []  # 학교명을 저장할 리스트

            single_mode = self.mode_var.get() == "single"
            combined_data = []
            combined_school_name = "단일학교" if single_mode else None

            # 교과(군) 모집 데이터 불러오기
            subject_group_mapping = self.load_subject_group_mapping("subject_group_mapping.json")
            self.add_log("교과(군) 매핑 데이터를 불러왔습니다.")

            # 선택된 모든 파일 처리
            for file_path in self.file_paths:
                # 파일명에서 괄호 안의 내용을 추출
                filename = os.path.basename(file_path)
                school_match = re.search(r'\((.*?)\)', filename)
                school_name = school_match.group(1) if school_match else "알수없음"
                # 학교명 수정 ('xx고'를 'xx고등학교'로)
                if school_name.endswith('고'):
                    school_name = school_name[:-1] + '고등학교'
                if single_mode:
                    school_name = "단일학교"

                school_names.append(school_name)

                self.add_log(f"파일 로드 중: {filename} (학교명: {school_name})")
                wb = openpyxl.load_workbook(file_path, data_only=True)
                results = self.process_workbook(wb)
                wb.close()

                if single_mode:
                    combined_data.extend(results)
                else:
                    school_data.append({
                        'school_name': school_name,
                        'data': results
                    })

            if single_mode and combined_data:
                output_file = os.path.join(os.getcwd(), "결과집계표.xlsx")
                school_data = [{
                    'school_name': combined_school_name or '단일학교',
                    'data': combined_data
                }]
                self.save_results(school_data, output_file, subject_group_mapping, school_names, single_school=True)
                self.add_log("결과 파일이 저장되었습니다.")
            elif not single_mode and school_data:
                output_file = os.path.join(os.getcwd(), "결과집계표.xlsx")
                self.save_results(school_data, output_file, subject_group_mapping, school_names, single_school=False)
                self.add_log("결과 파일이 저장되었습니다.")
            else:
                self.add_log("처리할 데이터가 없습니다.")
                messagebox.showwarning("경고", "처리할 데이터가 없습니다.")
                
        except Exception as e:
            error_msg = f"오류 발생: {str(e)}"
            self.add_log(error_msg)
            messagebox.showerror("Error", error_msg)
    
    def save_results(self, school_data, output_path, subject_group_mapping, school_names, single_school=False):
        wb = openpyxl.Workbook()
        
        # 첫 번째 시트: 교사별 시수 현황 (수정됨 - get_subject_group 사용)
        ws1 = wb.active
        ws1.title = "교사별시수현황"
        
        headers = ['학교명', '교사명', '과목', '총시수', '교과(군)']
        for col, header in enumerate(headers, 1):
            ws1.cell(row=1, column=col, value=header)

        # 데이터를 교사별로 정리
        current_row = 2
        for school in school_data:
            school_name = school['school_name']
            data = school['data']
            teacher_data = {}
            for item in data:
                teacher = item['교사명']
                if teacher not in teacher_data:
                    teacher_data[teacher] = []
                teacher_data[teacher].append(item)
            
            for teacher in sorted(teacher_data.keys()):
                subject_groups = set(
                    self.get_subject_group(item['과목'], subject_group_mapping)  # 수정된 부분
                    for item in teacher_data[teacher]
                )
                
                filtered_groups = self.filter_subject_groups(subject_groups)
                subject_groups_str = ', '.join(filtered_groups)
                for item in teacher_data[teacher]:
                    ws1.cell(row=current_row, column=1, value=school_name)
                    ws1.cell(row=current_row, column=2, value=teacher)
                    ws1.cell(row=current_row, column=3, value=item['과목'])
                    ws1.cell(row=current_row, column=4, value=item['총시수'])
                    subject_group = self.get_subject_group(item['과목'], subject_group_mapping)  # 수정된 부분
                    ws1.cell(row=current_row, column=5, value=subject_group)
                    current_row += 1
        self.autofit_columns(ws1)
        # 과목명이 긴 경우가 많아 C열 너비를 다시 계산해 자동 맞춤
        max_len_c = 0
        for row in range(1, ws1.max_row + 1):
            value = ws1.cell(row=row, column=3).value
            if value is not None:
                max_len_c = max(max_len_c, len(str(value)))
        ws1.column_dimensions['C'].width = max(max_len_c + 2, 10)

        # 두 번째 시트: 교사별 총계 (수정됨 - get_subject_group 사용)
        ws2 = wb.create_sheet(title="교사별총시수")
        
        summary_headers = ['학교명', '교사명', '담당교과', '총시수', '담당과목 수', '담당과목명', '교과(군)조합']
        for col, header in enumerate(summary_headers, 1):
            ws2.cell(row=1, column=col, value=header)

        # 학교별 데이터를 통합하기 위한 딕셔너리
        merged_teacher_data = {}
        
        # 모든 학교 데이터를 순회하면서 같은 학교의 같은 교사 데이터 통합
        for school in school_data:
            school_name = school['school_name']
            data = school['data']
            
            for item in data:
                teacher = item['교사명']
                key = (school_name, teacher)  # 학교명과 교사명으로 키 생성
                
                if key not in merged_teacher_data:
                    merged_teacher_data[key] = []
                merged_teacher_data[key].append(item)

        # 통합된 데이터를 시트에 작성
        current_row = 2
        total_all_teachers = 0
        total_all_hours = 0
        
        # 학교별로 정렬하여 데이터 작성
        for (school_name, teacher), items in sorted(merged_teacher_data.items()):
            # 담당교과(군) 추출 및 정렬 (수정된 부분)
            subject_groups = set(
                self.get_subject_group(item['과목'], subject_group_mapping)  # 수정된 부분
                for item in items
            )
            
            # 규칙 적용하여 교과 필터링
            filtered_groups = self.filter_subject_groups(subject_groups)
            subject_groups_str = ', '.join(filtered_groups)

            # 과목명 목록 생성 (중복 제거)
            subject_names = sorted(set(item['과목'] for item in items))
            subject_names_str = ', '.join(subject_names)

            # 총 시수 계산
            total_hours = sum(item['총시수'] for item in items)
            subject_count = len(set(item['과목'] for item in items))  # 중복 제거된 과목 수

            # 교과(군) 조합 문자열 생성 (수정된 부분)
            original_groups = sorted(set(
                self.get_subject_group(item['과목'], subject_group_mapping)  # 수정된 부분
                for item in items
            ))
            combination_str = ' + '.join(original_groups) if len(original_groups) >= 2 else ""

            # 데이터 입력
            ws2.cell(row=current_row, column=1, value=school_name)
            ws2.cell(row=current_row, column=2, value=teacher)
            ws2.cell(row=current_row, column=3, value=subject_groups_str)
            ws2.cell(row=current_row, column=4, value=total_hours)
            ws2.cell(row=current_row, column=5, value=subject_count)
            ws2.cell(row=current_row, column=6, value=subject_names_str)
            ws2.cell(row=current_row, column=7, value=combination_str)
            
            current_row += 1
            total_all_hours += total_hours

        # 학교별 교사 수 계산 (중복 제거)
        unique_teachers = len(set((school_name, teacher) for school_name, teacher in merged_teacher_data.keys()))
        
        # 전체 총계 추가
        ws2.cell(row=current_row, column=2, value="전체 교사수")
        ws2.cell(row=current_row, column=4, value=unique_teachers)
        current_row += 1
        ws2.cell(row=current_row, column=2, value="전체 시수")
        ws2.cell(row=current_row, column=4, value=total_all_hours)
                
        self.autofit_columns(ws2)
        
        # 세 번째 시트: 학교통계 (수정됨 - get_subject_group 사용)
        ws3 = wb.create_sheet(title="학교통계")
        
        # 헤더 생성
        headers = ['학교명']
        header_row = 1
        
        # 헤더 동적 생성을 위한 최대값 계산
        max_subjects = 1
        max_groups = 1
        aggregated_teacher_subject_counts = {}
        total_teachers_all = 0
        for school in school_data:
            data = school['data']
            teacher_data = {}
            for item in data:
                teacher = item['교사명']
                if teacher not in teacher_data:
                    teacher_data[teacher] = []
                teacher_data[teacher].append(item)
            
            # max_subjects 계산 수정
            if teacher_data.values():  # 값이 있는 경우에만 max 계산
                current_max = max(len(subjects) for subjects in teacher_data.values())
                max_subjects = max(max_subjects, current_max)

            teacher_subject_groups = {}
            for item in data:
                teacher = item['교사명']
                subject_group = self.get_subject_group(item['과목'], subject_group_mapping)  # 수정된 부분
                if teacher not in teacher_subject_groups:
                    teacher_subject_groups[teacher] = set()
                teacher_subject_groups[teacher].add(subject_group)
            
            # max_groups 계산 수정
            if teacher_subject_groups.values():  # 값이 있는 경우에만 max 계산
                current_max = max(len(groups) for groups in teacher_subject_groups.values())
                max_groups = max(max_groups, current_max)   
        
        # 1. 다과목지도 현황 헤더
        for i in range(1, max_subjects + 1):
            headers.extend([
                f'{i}과목_교사수',
                f'{i}과목_비율'
            ])
        
        # 2. 교과(군)별 통계 헤더 생성을 위한 모든 교과(군) 수집 (수정된 부분)
        all_subject_groups = set()
        for school in school_data:
            data = school['data']
            for item in data:
                subject_group = self.get_subject_group(item['과목'], subject_group_mapping)  # 수정된 부분
                all_subject_groups.add(subject_group)
        
        # 교과(군)별 통계 헤더
        subject_group_column_map = {}
        # 전문적인 파스텔 톤 색상 팔레트 적용
        color_palette = [
            'AED6F1', 'A9DFBF', 'F9E79F', 'F5CBA7',
            'D2B4DE', 'A9CCE3', 'FADBD8', 'D7BDE2'
        ]
        for idx, group in enumerate(sorted(all_subject_groups)):
            start_idx = len(headers) + 1
            headers.extend([
                f'{group}_교과(군)_교사수',
                f'{group}_교과(군)_교사의_총시수',
                f'{group}_교과(군)_과목의_총시수',
                f'{group}_교과(군)_교사의_평균시수',
                f'{group}_교과(군)_교사의_평균과목수'
            ])
            subject_group_column_map[group] = list(range(start_idx, start_idx + 5))
        subject_group_colors = {g: color_palette[i % len(color_palette)] for i, g in enumerate(sorted(all_subject_groups))}

        # 3. 복수 교과(군) 통계 헤더
        for i in range(1, max_groups + 1):
            headers.extend([
                f'{i}개교과군_교사수',
                f'{i}개교과군_비율'
            ])
        
        # 4. 총계 관련 헤더 수정
        headers.extend([
            '전체_교사수',
            '전체_시수',
            '평균시수',  # 추가
            '전체_과목수',
            '개설_과목수',  # 추가
            '평균_과목수'
        ])
        
        # 헤더 쓰기 (이모지 아이콘 포함)
        emoji_map = {
            '학교명': '🏫',
            '전체_교사수': '👩\u200d🏫',
            '전체_시수': '⏱️',
            '평균시수': '📊',
            '전체_과목수': '📚',
            '개설_과목수': '🆕',
            '평균_과목수': '💡'
        }
        for col, header in enumerate(headers, 1):
            icon = emoji_map.get(header, '')
            display = f"{header} {icon}" if icon else header
            ws3.cell(row=1, column=col, value=display)
        
        # 데이터 입력 (각 학교별로) - 수정된 부분들이 포함됨
        current_row = 2
        for school in school_data:
            col = 1
            school_name = school['school_name']
            data = school['data']
            
            # 교사 데이터 정리
            teacher_data = {}
            for item in data:
                teacher = item['교사명']
                if teacher not in teacher_data:
                    teacher_data[teacher] = []
                teacher_data[teacher].append(item)
            
            total_teachers = len(teacher_data)
            total_teachers_all += total_teachers
            
            # 학교명 입력
            ws3.cell(row=current_row, column=col, value=school_name)
            col += 1
            
            # 1. 다과목지도 현황 데이터
            teacher_subject_counts = {}
            for teacher in teacher_data:
                subject_count = len(teacher_data[teacher])
                teacher_subject_counts[subject_count] = teacher_subject_counts.get(subject_count, 0) + 1
                aggregated_teacher_subject_counts[subject_count] = aggregated_teacher_subject_counts.get(subject_count, 0) + 1
            
            for i in range(1, max_subjects + 1):
                count = teacher_subject_counts.get(i, 0)
                percentage = round((count / total_teachers) * 100, 1) if total_teachers > 0 else 0
                ws3.cell(row=current_row, column=col, value=count)
                ws3.cell(row=current_row, column=col+1, value=percentage)
                col += 2
            
            # 2. 교과(군)별 통계 데이터 (수정된 부분)
            subject_group_stats = {}
            for item in data:
                subject_group = self.get_subject_group(item['과목'], subject_group_mapping)  # 수정된 부분
                teacher = item['교사명']
                hours = item['총시수']
                
                if subject_group not in subject_group_stats:
                    subject_group_stats[subject_group] = {'teachers': set(), 'total_hours': 0}
                
                subject_group_stats[subject_group]['teachers'].add(teacher)
                subject_group_stats[subject_group]['total_hours'] += hours
            
            # 교과(군)별 통계 데이터 입력 부분 수정
            for group in sorted(all_subject_groups):
                stats = subject_group_stats.get(group, {'teachers': set(), 'total_hours': 0})
                teacher_count = len(stats['teachers'])
                total_hours = stats['total_hours']
                
                # 교과(군)별 평균과목수 계산 부분 수정
                group_subjects = []  # set 대신 list 사용
                for teacher in stats['teachers']:
                    teacher_subjects = set()  # 각 교사가 담당하는 해당 교과(군)의 과목들
                    for item in data:
                        if item['교사명'] == teacher:
                            if self.get_subject_group(item['과목'], subject_group_mapping) == group:  # 수정된 부분
                                teacher_subjects.add(item['과목'])
                    group_subjects.append(len(teacher_subjects))  # add 대신 append 사용

                # 평균 과목 수 계산 - 전체 교사의 과목 수 합계를 교사 수로 나눔
                avg_subjects = round(sum(group_subjects) / teacher_count, 2) if teacher_count > 0 else 0
                
                # 수식 설정을 위해 셀 참조 구하기
                count_cell = ws3.cell(row=current_row, column=col)
                teacher_total_cell = ws3.cell(row=current_row, column=col + 1)
                subject_total_cell = ws3.cell(row=current_row, column=col + 2)
                avg_cell = ws3.cell(row=current_row, column=col + 3)

                if single_school:
                    count_cell.value = f"=COUNTIF('교사별총시수'!$C:$C,\"*{group}*\")"
                    teacher_total_cell.value = (
                        f"=SUMIFS('교사별총시수'!$D:$D,"
                        f"'교사별총시수'!$C:$C,\"*{group}*\")"
                    )
                    subject_total_cell.value = (
                        f"=SUMIFS('교사별시수현황'!$D:$D,"
                        f"'교사별시수현황'!$E:$E,\"{group}\")"
                    )
                    avg_cell.value = (
                        f"=IFERROR(AVERAGEIF('교사별총시수'!$C:$C,\"*{group}*\",'교사별총시수'!$D:$D),0)"
                    )
                else:
                    school_ref = f"$A{current_row}"
                    count_cell.value = (
                        f"=COUNTIFS('교사별총시수'!$A:$A,{school_ref},"
                        f"'교사별총시수'!$C:$C,\"*{group}*\")"
                    )
                    teacher_total_cell.value = (
                        f"=SUMIFS('교사별총시수'!$D:$D,"
                        f"'교사별총시수'!$A:$A,{school_ref},"
                        f"'교사별총시수'!$C:$C,\"*{group}*\")"
                    )
                    subject_total_cell.value = (
                        f"=SUMIFS('교사별시수현황'!$D:$D,"
                        f"'교사별시수현황'!$A:$A,{school_ref},"
                        f"'교사별시수현황'!$E:$E,\"{group}\")"
                    )
                    avg_cell.value = (
                        f"=IFERROR(AVERAGEIFS('교사별총시수'!$D:$D,"
                        f"'교사별총시수'!$A:$A,{school_ref},"
                        f"'교사별총시수'!$C:$C,\"*{group}*\"),0)"
                    )
                ws3.cell(row=current_row, column=col + 4, value=avg_subjects)  # 평균과목수 입력
                col += 5  # 컬럼 개수 5로 수정
            
            # 3. 복수 교과(군) 통계 데이터 (수정된 부분)
            teacher_subject_groups = {}
            for item in data:
                teacher = item['교사명']
                subject_group = self.get_subject_group(item['과목'], subject_group_mapping)  # 수정된 부분
                if teacher not in teacher_subject_groups:
                    teacher_subject_groups[teacher] = set()
                teacher_subject_groups[teacher].add(subject_group)
            
            multi_group_stats = {}
            for teacher, groups in teacher_subject_groups.items():
                group_count = len(groups)
                multi_group_stats[group_count] = multi_group_stats.get(group_count, 0) + 1
            
            for i in range(1, max_groups + 1):
                count = multi_group_stats.get(i, 0)
                percentage = round((count / total_teachers) * 100, 1) if total_teachers > 0 else 0
                ws3.cell(row=current_row, column=col, value=count)
                ws3.cell(row=current_row, column=col+1, value=percentage)
                col += 2
            
            # 4. 총계 데이터
            total_teachers = len(teacher_data)
            total_hours = sum(item['총시수'] for item in data)
            avg_hours = round(total_hours / total_teachers, 2) if total_teachers > 0 else 0  # 평균시수 계산
            
            total_subjects = sum(count * subjects for subjects, count in teacher_subject_counts.items())
            unique_subjects = len(set(item['과목'] for item in data))  # 중복 제거한 과목 수
            avg_subjects = round(total_subjects / total_teachers, 2) if total_teachers > 0 else 0
            
            ws3.cell(row=current_row, column=col, value=total_teachers)
            col += 1
            ws3.cell(row=current_row, column=col, value=total_hours)
            col += 1
            ws3.cell(row=current_row, column=col, value=avg_hours)  # 평균시수 입력
            col += 1
            ws3.cell(row=current_row, column=col, value=total_subjects)
            col += 1
            ws3.cell(row=current_row, column=col, value=unique_subjects)  # 개설과목수 입력
            col += 1
            ws3.cell(row=current_row, column=col, value=avg_subjects)
            current_row += 1

        data_end_row = current_row - 1

        if not single_school:
            teacher_col = headers.index('전체_교사수') + 1
            avg_col = headers.index('평균시수') + 1
            cats = Reference(ws3, min_col=1, min_row=2, max_row=data_end_row)
            data_ref = Reference(ws3, min_col=teacher_col, max_col=avg_col, min_row=1, max_row=data_end_row)
            bar_chart = BarChart()
            bar_chart.title = "교사수 및 평균시수"
            bar_chart.add_data(data_ref, titles_from_data=True)
            bar_chart.set_categories(cats)
            ws3.add_chart(bar_chart, f"A{data_end_row + 3}")

            summary_start = data_end_row + 2
            ws3.cell(row=summary_start, column=1, value="과목수")
            ws3.cell(row=summary_start, column=2, value="비율")
            for i in range(1, max_subjects + 1):
                percent = round((aggregated_teacher_subject_counts.get(i, 0) / total_teachers_all) * 100, 2) if total_teachers_all > 0 else 0
                ws3.cell(row=summary_start + i, column=1, value=f"{i}과목")
                ws3.cell(row=summary_start + i, column=2, value=percent)
            pie = PieChart()
            labels = Reference(ws3, min_col=1, min_row=summary_start + 1, max_row=summary_start + max_subjects)
            data = Reference(ws3, min_col=2, min_row=summary_start, max_row=summary_start + max_subjects)
            pie.add_data(data, titles_from_data=True)
            pie.set_categories(labels)
            pie.title = "과목수별 비율"
            ws3.add_chart(pie, f"A{summary_start + max_subjects + 2}")

        # 단일 학교 모드인 경우 세로 형태로 변환
        if single_school and len(school_data) == 1:
            headers_row = [ws3.cell(row=1, column=i).value for i in range(1, ws3.max_column + 1)]
            values_row = [ws3.cell(row=2, column=i).value for i in range(1, ws3.max_column + 1)]
            ws3.delete_rows(1, ws3.max_row)
            ws3.cell(row=1, column=1, value="항목")
            ws3.cell(row=1, column=2, value="값")

            for idx, (h, v) in enumerate(zip(headers_row, values_row), start=2):
                ws3.cell(row=idx, column=1, value=h)
                ws3.cell(row=idx, column=2, value=v)

            # ----- Charts for single school -----
            # 교과(군)별 평균시수 막대그래프 데이터 수집
            avg_rows = []
            for r in range(2, ws3.max_row + 1):
                label = str(ws3.cell(row=r, column=1).value)
                if label.endswith('_교과(군)_교사의_평균시수'):
                    subject = label.split('_')[0]
                    avg_rows.append((subject, r))

            chart_start = ws3.max_row + 2
            for idx, (subject, row) in enumerate(avg_rows, start=0):
                ws3.cell(row=chart_start + idx, column=1, value=subject)
                ws3.cell(row=chart_start + idx, column=2, value=ws3.cell(row=row, column=2).value)

            if avg_rows:
                cats = Reference(ws3, min_col=1, min_row=chart_start, max_row=chart_start + len(avg_rows) - 1)
                data = Reference(ws3, min_col=2, min_row=chart_start, max_row=chart_start + len(avg_rows) - 1)
                bar = BarChart()
                bar.title = "교과(군)별 평균시수"
                bar.add_data(data, titles_from_data=False)
                bar.set_categories(cats)
                bar.legend = None
                bar.dataLabels = DataLabelList()
                bar.dataLabels.showVal = True
                bar.width = 12
                bar.height = 8
                ws3.add_chart(bar, "D2")

            # n과목 교사 비율 도넛차트 데이터 수집
            ratio_rows = []
            for r in range(2, chart_start):
                label = str(ws3.cell(row=r, column=1).value)
                if label.endswith('과목_비율') and not label.endswith('개교과군_비율'):
                    ratio_rows.append((label.split('_')[0], r))

            donut_start = chart_start + len(avg_rows) + 1
            for idx, (label, row) in enumerate(ratio_rows, start=0):
                ws3.cell(row=donut_start + idx, column=1, value=label)
                ws3.cell(row=donut_start + idx, column=2, value=ws3.cell(row=row, column=2).value)

            if ratio_rows:
                cats = Reference(ws3, min_col=1, min_row=donut_start, max_row=donut_start + len(ratio_rows) - 1)
                data = Reference(ws3, min_col=2, min_row=donut_start, max_row=donut_start + len(ratio_rows) - 1)
                donut = DoughnutChart()
                donut.style = 3  # 차트 디자인 3 적용
                donut.title = "과목수별 비율"
                donut.add_data(data, titles_from_data=False)
                donut.set_categories(cats)
                donut.dataLabels = DataLabelList()
                donut.dataLabels.showPercent = True
                donut.width = 12
                donut.height = 8
                ws3.add_chart(donut, "L2")

        # 스타일 적용
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        # 대시보드 스타일 헤더용 그라데이션 배경 설정
        header_fill = GradientFill(stop=('DCEFFB', 'E8DAEF'))
        header_font = Font(bold=True)
        
        # 모든 시트에 스타일 적용
        for ws in [ws1, ws2, ws3]:
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    if cell.row == 1:  # 헤더 행
                        cell.font = header_font
                        cell.fill = header_fill

        # 교과별 색상 적용 및 평균시수 서식 지정
        if single_school and len(school_data) == 1:
            # 세로 레이아웃에 맞게 색상 및 서식 적용
            for row in range(2, ws3.max_row + 1):
                label = str(ws3.cell(row=row, column=1).value)
                value_cell = ws3.cell(row=row, column=2)
                for group, color in subject_group_colors.items():
                    if label.startswith(group):
                        fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                        ws3.cell(row=row, column=1).fill = fill
                        value_cell.fill = fill
                if '평균시수' in label:
                    value_cell.number_format = '0.00'
        else:
            for group, cols in subject_group_column_map.items():
                fill = PatternFill(start_color=subject_group_colors[group], end_color=subject_group_colors[group], fill_type='solid')
                avg_col = cols[3]
                for row in range(1, data_end_row + 1):
                    for col_idx in cols:
                        ws3.cell(row=row, column=col_idx).fill = fill
                    if row >= 2:
                        ws3.cell(row=row, column=avg_col).number_format = '0.00'

            avg_hours_col = headers.index('평균시수') + 1
            for row in range(2, data_end_row + 1):
                ws3.cell(row=row, column=avg_hours_col).number_format = '0.00'

            # 조건부 서식: 데이터 바 및 컬러 스케일 적용
            teacher_col_letter = get_column_letter(teacher_col)
            avg_hours_col_letter = get_column_letter(avg_hours_col)
            ws3.conditional_formatting.add(
                f"{teacher_col_letter}2:{teacher_col_letter}{data_end_row}",
                DataBarRule(start_type='num', start_value=0, end_type='max', color='5DADE2')
            )
            ws3.conditional_formatting.add(
                f"{avg_hours_col_letter}2:{avg_hours_col_letter}{data_end_row}",
                ColorScaleRule(start_type='min', start_color='FFFFFF', mid_type='percentile', mid_value=50,
                               mid_color='FFF5CC', end_type='max', end_color='A9DFBF')
            )

            # KPI 카드 작성
            kpi_start = ws3.max_row + 2
            kpis = [
                ('학교수', len(school_data), '🏫'),
                ('총 교사수', total_teachers_all, '👩\u200d🏫'),
                ('총 시수', total_all_hours, '⏱️'),
                (
                    '평균 시수',
                    round(total_all_hours / total_teachers_all, 2) if total_teachers_all else 0,
                    '📊'
                )
            ]
            card_width = 4
            for idx, (label, value, icon) in enumerate(kpis):
                start_col = 1 + idx * card_width
                end_col = start_col + card_width - 1
                ws3.merge_cells(start_row=kpi_start, start_column=start_col,
                                end_row=kpi_start, end_column=end_col)
                cell = ws3.cell(row=kpi_start, column=start_col, value=f"{icon} {label}: {value}")
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = Font(bold=True)
                cell.fill = GradientFill(stop=('FFFFFF', 'D6EAF8'))
            ws3.row_dimensions[kpi_start].height = 25

        # 수식 셀 색상 지정
        formula_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
        for row_cells in ws3.iter_rows(min_row=2, max_row=ws3.max_row):
            for c in row_cells:
                if c.data_type == 'f':
                    c.fill = formula_fill

        # 열 너비 자동 조정 (학교통계 시트)
        self.autofit_columns(ws3)
        # 학교통계 시트의 A열은 학교명이 길어질 수 있어 넉넉하게 설정
        ws3.column_dimensions['A'].width = 40
        # B열은 비고 등의 짧은 값을 담으므로 고정 폭 지정
        ws3.column_dimensions['B'].width = 20
        
        # 네 번째 시트: 복수 교과(군) 조합 현황 (수정된 부분)
        ws4 = wb.create_sheet(title="교과군조합현황")
        
        # 헤더 설정
        combination_headers = ['학교명', '교과(군) 조합', '교사수', '해당 교사명']
        for col, header in enumerate(combination_headers, 1):
            ws4.cell(row=1, column=col, value=header)
        
        current_row = 2
        for school in school_data:
            school_name = school['school_name']
            data = school['data']
            
            # 교사별 담당 교과(군) 수집 (수정된 부분)
            teacher_subject_groups = {}
            for item in data:
                teacher = item['교사명']
                subject_group = self.get_subject_group(item['과목'], subject_group_mapping)  # 수정된 부분
                
                if teacher not in teacher_subject_groups:
                    teacher_subject_groups[teacher] = set()
                teacher_subject_groups[teacher].add(subject_group)
            
            # 교과(군) 조합별 교사 수집
            group_combinations = {}
            for teacher, groups in teacher_subject_groups.items():
                if len(groups) >= 2:  # 2개 이상의 교과(군)을 담당하는 경우
                    groups_tuple = tuple(sorted(groups))  # 정렬하여 동일한 조합을 같은 것으로 처리
                    if groups_tuple not in group_combinations:
                        group_combinations[groups_tuple] = []
                    group_combinations[groups_tuple].append(teacher)
            
            # 조합별 데이터 입력
            for groups_tuple, teachers in sorted(group_combinations.items(), key=lambda x: (-len(x[0]), x[0])):
                combination_str = ' + '.join(groups_tuple)
                teacher_names = ', '.join(sorted(teachers))
                
                ws4.cell(row=current_row, column=1, value=school_name)
                ws4.cell(row=current_row, column=2, value=combination_str)
                ws4.cell(row=current_row, column=3, value=len(teachers))
                ws4.cell(row=current_row, column=4, value=teacher_names)
                current_row += 1
        
        # 스타일 적용
        for row in range(1, current_row):
            for col in range(1, len(combination_headers) + 1):
                cell = ws4.cell(row=row, column=col)
                cell.border = thin_border
                if row == 1:  # 헤더 행
                    cell.font = header_font
                    cell.fill = header_fill
                # 교사명 열은 왼쪽 정렬, 나머지는 가운데 정렬
                if col == 4:
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 열 너비 자동 조정
        self.autofit_columns(ws4)
        
        # 눈금선 숨기기
        ws4.sheet_view.showGridLines = False    
        # 모든 시트 눈금선 숨기기
        for ws in [ws1, ws2, ws3]:
            ws.sheet_view.showGridLines = False
        
        # 엑셀 파일 저장
        try:
            wb.save(output_path)
            self.open_file(output_path)
        except PermissionError:
            base, ext = os.path.splitext(output_path)
            count = 2
            while True:
                new_output_path = f"{base}({count}){ext}"
                try:
                    wb.save(new_output_path)
                    self.open_file(new_output_path)
                    break
                except PermissionError:
                    count += 1

    def run(self):
        self.root.mainloop()

if __name__ == "__main__":
    app = TimeTableProcessor()
    app.run()